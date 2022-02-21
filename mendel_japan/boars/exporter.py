"""ダウンロード用のExcelファイルを作成してレスポンスとして返す"""

import flask
import flask_sqlalchemy
import pandas as pd
import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import os
from datetime import datetime


from config import engine
from mendel_japan.models import Farm, Line


def downloadExcel(
        boars_query: flask_sqlalchemy.BaseQuery) -> flask.wrappers.Response:
    """ダウンロード用のExcelファイルを作成してレスポンスとして返す

    ・選択した条件の雄(boars_query)をデータフレームに変換
    ・カラム名を日本語に変換
    ・リレーションしている項目を変換
    ・Excelファイルに出力
    ・Excelファイルをレスポンスとして返す

    Args:
        boars_query (flask_sqlalchemy.BaseQuery): 選択した条件(SQL)

    Returns:
        flask.wrappers.Response: Excelファイルのレスポンス
    """
    boars: pd.DataFrame = pd.read_sql(boars_query.statement, con=engine)
    boars_rename: pd.DataFrame = data_rename(boars)
    file_name: str = add_workbook(boars_rename)
    return add_response(file_name)


def data_rename(boars: pd.DataFrame) -> pd.DataFrame:
    """カラム名とリレーションしている項目を変換して返す

    ・カラム名を日本語に変換
    ・農場カラムの内容をFarm.idから農場名に変換
    ・系統カラムの内容をLine.idから系統(略)に変換

    Args:
        boars (pd.DataFrame): 未加工の雄一覧

    Returns:
        pd.DataFrame: 変換後の雄一覧
    """
    boars_columns_rename: pd.DataFrame = boars.rename(columns=rename_column())
    boars_columns_rename['農場'] = \
        boars_columns_rename.農場.map(lambda x: Farm.query.get(x).name)
    boars_columns_rename['系統'] = \
        boars_columns_rename.系統.map(lambda x: Line.query.get(x).abbreviation)
    return boars_columns_rename[rename_column().values()]


def rename_column() -> dict:
    """カラム名を日本語に変換する前後を辞書型で返す

    Returns:
        dict: 変換前後のカラム名
    """
    return {
        'farm_id': '農場',
        'tattoo': 'タトゥー',
        'name': '雄ID',
        'line_id': '系統',
        'birth_on': '生年月日',
        'culling_on': '淘汰日',
    }


def add_workbook(boars: pd.DataFrame) -> str:
    """Excelファイルを新規作成してファイル名を返す

    ・Excelファイルを新規作成
    ・データフレームの内容を入力
    ・フォーマットを整えて保存しファイル名を返す

    Args:
        boars (pd.DataFrame): 雄一覧

    Returns:
        str: ファイル名
    """
    wb: xl.Workbook = xl.Workbook()
    ws: xl.worksheet = wb.active

    input_worksheet(ws, boars)
    add_format(ws)

    now: datetime = datetime.now().strftime('%y%m%d%H%M%S')
    file_name: str = f'{now}_boar_list.xlsx'
    wb.save(file_name)
    wb.close()
    return file_name


def input_worksheet(ws: xl.worksheet, boars: pd.DataFrame) -> None:
    """データフレームの内容をExcelファイルに入力

    Args:
        ws (xl.worksheet): ワークシート
        boars (pd.DataFrame): 雄一覧
    """
    for col, title in enumerate(boars.columns.values, 1):
        ws.cell(1, col).value = title

    for row, boar in enumerate(boars.itertuples(), 2):
        for col, data in enumerate(boar[1:], 1):
            ws.cell(row, col).value = data


def add_format(ws: xl.worksheet) -> None:
    """フォーマットの設定

    Args:
        ws (xl.worksheet): ワークシート
    """
    for col in ws.columns:
        change_font(ws, col, 12)
        change_width(ws, col)
        background_color(col)
        add_border(col)


def change_font(ws: xl.worksheet, col: tuple, size: int) -> None:
    """フォントの変更と縮小して全体を表示設定

    Args:
        ws (xl.worksheet): ワークシート
        col (xl.column): 列オブジェクト
        size (int, optional): 文字サイズ
    """
    for cell in col:
        font: xl.styles = xl.styles.fonts.Font(name='Yu Gothic', size=size)
        ws[cell.coordinate].font = font
        ws[cell.coordinate].alignment = Alignment(shrinkToFit=True)


def change_width(ws: xl.worksheet, col: tuple) -> None:
    """列幅の設定

    Args:
        ws (xl.worksheet): ワークシート
        col (xl.column): 列オブジェクト
    """
    column_initial: str = col[0].column_letter
    ws.column_dimensions[column_initial].width = 13


def background_color(col: tuple) -> None:
    """背景色の変更

    ・1行目のみ水色

    Args:
        col (xl.column): 列オブジェクト
    """
    for cell in col:
        if cell.row == 1:
            fill(cell, 'CCECFF')


def fill(cell: xl.cell, color: str) -> None:
    """背景色を指定の色に変更

    Args:
        cell (xl.cell): セルオブジェクト
        color (str): 色コード
    """
    cell.fill = xl.styles.PatternFill(
        patternType='solid', fgColor=color, bgColor=color)


def add_border(col: tuple) -> None:
    """セルに罫線を設定する

    四方を通常の罫線で囲む

    Args:
        ws (xl.worksheet): ワークシート
        col (xl.column): 列オブジェクト
    """
    for cell in col:
        side: xl.styles = Side(style='thin', color='000000')
        border: xl.styles = \
            Border(top=side, bottom=side, left=side, right=side)
        cell.border = border


def add_response(file_name: str) -> flask.wrappers.Response:
    """レスポンスを返す

    中段は何の設定か良くわからない

    Args:
        file_name (str): ファイル名

    Returns:
        flask.wrappers.Response: レスポンス(Excelファイル)
    """
    response: flask.wrappers.Response = flask.make_response()
    wb = open(file_name, "rb")
    response.data = wb.read()
    wb.close()

    response.headers["Content-Disposition"] = \
        "attachment; filename=" + file_name
    XLSX_MIMETYPE = \
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.mimetype = XLSX_MIMETYPE

    os.remove(file_name)
    return response
